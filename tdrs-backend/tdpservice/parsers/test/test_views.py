"""Tests for the views in the parsers app."""

import pytest
import base64
import openpyxl
from .. import parse
from tdpservice.parsers.views import ParsingErrorViewSet
from .. import util

@pytest.fixture
def test_datafile(stt_user, stt):
    """Fixture for small_incorrect_file_cross_validator."""
    return util.create_test_datafile('small_incorrect_file_cross_validator.txt', stt_user, stt)


@pytest.mark.django_db
def test_parsing_error_viewset_list(client, mocker, test_datafile):
    """Test the django rest framework parsing error viewset list."""
    # parse datafile
    test_datafile.year = 2021
    test_datafile.quarter = 'Q2'
    parse.parse_datafile(test_datafile)

    id = test_datafile.id
    view = ParsingErrorViewSet()
    view.format_kwarg = None
    request = mocker.Mock()

    request.query_params = {'file': id}
    view.request = request

    ser = view.list(request).data['xls_report']

    decoded_ser = base64.b64decode(ser)
    # write the excel file to disk
    with open('mycls.xlsx', 'wb') as f:
        f.write(decoded_ser)

    # read the excel file from disk
    wb = openpyxl.load_workbook('mycls.xlsx')
    ws = wb.get_sheet_by_name('Sheet1')

    COL_ERROR_MESSAGE = 5

    assert ws.cell(row=1, column=1).value == "Error reporting in TDP is still in development.We'll" \
        + " be in touch when it's ready to use!For now please refer to the reports you receive via email"
    assert ws.cell(row=4, column=COL_ERROR_MESSAGE).value == "if cash amount :873 validator1 passed" \
        + " then number of months 0 is not larger than 0."


@pytest.mark.django_db
def test_parsing_error_viewset_list_no_fields_json(mocker, test_datafile):
    """Test the django rest framework parsing error viewset list."""
    # parse datafile
    test_datafile.year = 2021
    test_datafile.quarter = 'Q2'
    parse.parse_datafile(test_datafile)

    # set fields_json to None
    from tdpservice.parsers.models import ParserError
    for error in ParserError.objects.all():
        error.fields_json = None
        error.save()

    id = test_datafile.id
    view = ParsingErrorViewSet()
    view.format_kwarg = None
    request = mocker.Mock()

    request.query_params = {'file': id}
    view.request = request

    ser = view.list(request).data['xls_report']

    decoded_ser = base64.b64decode(ser)
    # write the excel file to disk
    with open('mycls.xlsx', 'wb') as f:
        f.write(decoded_ser)

    # read the excel file from disk
    wb = openpyxl.load_workbook('mycls.xlsx')
    ws = wb.get_sheet_by_name('Sheet1')

    COL_ERROR_MESSAGE = 5

    assert ws.cell(row=1, column=1).value == "Error reporting in TDP is still in development.We'll" \
        + " be in touch when it's ready to use!For now please refer to the reports you receive via email"
    assert ws.cell(row=4, column=COL_ERROR_MESSAGE).value == "if CASH_AMOUNT :873 validator1 passed" \
        + " then NBR_MONTHS 0 is not larger than 0."

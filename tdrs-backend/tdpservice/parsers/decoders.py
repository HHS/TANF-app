"""Decoder and utility classes."""

from tdpservice.parsers.dataclasses import TupleRow, RawRow

from abc import ABC, abstractmethod
from enum import IntEnum, auto
import chardet
import csv
import logging
from openpyxl import load_workbook
import os
import puremagic

logger = logging.getLogger(__name__)


class Decoder(IntEnum):
    """Enum class for decoder types."""

    UTF8 = auto()
    CSV = auto()
    XLSX = auto()
    UNKNOWN = auto()


class BaseDecoder(ABC):
    """Abstract base class for all decoders."""

    def __init__(self, raw_file):
        super().__init__()
        self.raw_file = raw_file
        self.current_row_num = 0

        # Always ensure our file pointer is at the start
        self.raw_file.seek(0)

    @abstractmethod
    def get_record_type(self, raw_data) -> str:
        """To be implemented in child class."""
        pass

    @abstractmethod
    def get_header(self) -> RawRow:
        """To be implemented in child class."""
        pass

    @abstractmethod
    def decode(self) -> RawRow:
        """To be implemented in child class."""
        pass


class Utf8Decoder(BaseDecoder):
    """Decoder for UTF-8 files."""

    def get_record_type(self, raw_data):
        """Get the record type based on the raw data."""
        if raw_data.startswith('HEADER'):
            return "HEADER"
        elif raw_data.startswith('TRAILER'):
            return "TRAILER"
        else:
            return raw_data[0:2]

    def get_header(self):
        """Get the first line in the file. Assumed to be the header."""
        raw_data = self.raw_file.readline().decode().strip()
        return RawRow(data=raw_data, raw_len=len(raw_data), decoded_len=len(raw_data),
                      row_num=self.current_row_num, record_type="HEADER")

    def decode(self):
        """Decode and yield each row."""
        for raw_data in self.raw_file:
            self.current_row_num += 1
            raw_len = len(raw_data)
            raw_data = raw_data.decode().strip('\r\n')
            decoded_len = len(raw_data)
            record_type = self.get_record_type(raw_data)
            yield RawRow(data=raw_data, raw_len=raw_len, decoded_len=decoded_len,
                         row_num=self.current_row_num, record_type=record_type)


class CsvDecoder(BaseDecoder):
    """Decoder for csv files."""

    def __init__(self, raw_file):
        super().__init__(raw_file)
        self.local_file = None
        self.csv_file = None
        self._open_as_csv()

    def _open_as_csv(self):
        """Read binary csv to local storage and reopen in text mode."""
        name = self.raw_file.name.split('/')[-1]
        with open(f'/tmp/{name}', 'wb') as file:
            for line in self.raw_file:
                file.write(line)

        self.local_file = open(f'/tmp/{name}', 'rt')
        self.csv_file = csv.reader(self.local_file)

    def get_record_type(self, raw_data):
        """Get the record type based on the raw data."""
        # Until the need for more complicated logic arises, we assume this decoder is only being used for FRA files.
        return "TE1"

    def get_header(self):
        """Get the first line in the file. Assumed to be the header."""
        raw_data = None
        for line in self.csv_file:
            raw_data = line
            break
        # Very important to move pointer back to the begining since invoking the generator does not do it for us.
        self.local_file.seek(0)
        length = len(raw_data)
        return TupleRow(data=tuple(raw_data), raw_len=length, decoded_len=length,
                        row_num=self.current_row_num, record_type="HEADER")

    def decode(self):
        """Decode and yield each row."""
        for raw_data in self.csv_file:
            self.current_row_num += 1
            if not len(raw_data) or not any(raw_data) or str(raw_data[0]).startswith('#'):
                continue
            raw_len = len(raw_data)
            record_type = self.get_record_type(raw_data)
            yield TupleRow(data=tuple(raw_data), raw_len=raw_len, decoded_len=raw_len,
                           row_num=self.current_row_num, record_type=record_type)

    def __del__(self):
        """Close and delete the file when destructed."""
        try:
            self.local_file.close()
            if os.path.exists(self.local_file.name):
                os.remove(self.local_file.name)
                assert os.path.exists(self.local_file.name) is False
        except Exception:
            logger.exception("Encountered exception while closing and deleting file instance.")


class XlsxDecoder(BaseDecoder):
    """Decoder for xlsx files."""

    def __init__(self, raw_file):
        super().__init__(raw_file)
        self.work_book = load_workbook(raw_file)

    def get_record_type(self, raw_data):
        """Get the record type based on the raw data."""
        # Until the need for more complicated logic arises, we assume this decoder is only being used for FRA files.
        return "TE1"

    def get_header(self):
        """Get the first line in the file. Assumed to be the header."""
        for raw_data in self.work_book.active.iter_rows(values_only=True):
            length = len(raw_data)
            return TupleRow(data=raw_data, raw_len=length, decoded_len=length,
                            row_num=self.current_row_num, record_type="HEADER")

    def decode(self):
        """Decode and yield each row."""
        for raw_data in self.work_book.active.iter_rows(values_only=True):
            self.current_row_num += 1
            if not len(raw_data) or not any(raw_data) or str(raw_data[0]).startswith('#'):
                continue
            raw_len = len(raw_data)
            record_type = self.get_record_type(raw_data)
            yield TupleRow(data=raw_data, raw_len=raw_len, decoded_len=raw_len,
                           row_num=self.current_row_num, record_type=record_type)


class DecoderFactory:
    """Factory class to get/instantiate parsers."""

    @classmethod
    def get_suggested_decoder(cls, raw_file):
        """Try and determine what decoder to use based on file encoding and magic numbers."""
        # We need to guarantee that the file pointer is at the first byte
        raw_file.seek(0)
        extension = os.path.splitext(raw_file.name)[-1]

        # If our file has size zero, use the extension to try and determine the correct decoder. Default to UTF8 in
        # the worst case.
        if not len(raw_file):
            match extension:
                case ".csv":
                    logger.warning("File was empty. Returning CSV decoder based on extension.")
                    return Decoder.CSV
                case ".xlsx":
                    logger.warning("File was empty. Returning XLSX decoder based on extension.")
                    return Decoder.XLSX
                case _:
                    logger.warning("File was empty. Returning UTF8 decoder based on extension.")
                    return Decoder.UTF8

        data = raw_file.read(4096)
        char_result = chardet.detect(data)
        if char_result.get('encoding') == "ascii":
            confidence = char_result.get('confidence')
            if "csv" in os.path.splitext(raw_file.name)[-1]:
                logger.info(f"Returning CSV decoder with a confidence score of {confidence}")
                return Decoder.CSV
            logger.info(f"Returning UTF-8 decoder with a confidence score of {confidence}")
            return Decoder.UTF8
        else:
            try:
                predictions = puremagic.magic_string(data)
                most_confident = predictions[0]
                if most_confident.extension == "xlsx":
                    logger.info(f"Returning XLSX decoder with a confidence score of {most_confident.confidence}")
                    return Decoder.XLSX
            except puremagic.PureError:
                return Decoder.UNKNOWN
        return Decoder.UNKNOWN

    @classmethod
    def get_instance(cls, raw_file):
        """Return the correct parser class to be constructed manually."""
        decoder = cls.get_suggested_decoder(raw_file)
        match decoder:
            case Decoder.UTF8:
                return Utf8Decoder(raw_file)
            case Decoder.CSV:
                return CsvDecoder(raw_file)
            case Decoder.XLSX:
                return XlsxDecoder(raw_file)
            case Decoder.UNKNOWN:
                raise ValueError("Could not determine what decoder to use for file.")

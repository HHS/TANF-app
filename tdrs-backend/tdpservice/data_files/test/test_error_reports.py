"""Tests for generated data file error report workbooks."""

from io import BytesIO

from openpyxl import load_workbook
import pytest

from tdpservice.data_files.error_reports import ActiveClosedErrorReport
from tdpservice.data_files.models import DataFile
from tdpservice.data_files.test.factories import DataFileFactory
from tdpservice.data_files.parser_error_choices import ParserErrorCategoryChoices
from tdpservice.parsers.test.factories import ParserErrorFactory


KNOWLEDGE_CENTER_URL = (
    "https://tdp-project-updates.app.cloud.gov/knowledge-center/"
    "viewing-error-reports.html"
)


@pytest.fixture
def active_case_error_report_workbook():
    """Generate an Active Case Data error report workbook."""
    datafile = DataFileFactory(
        section=DataFile.Section.ACTIVE_CASE_DATA,
        program_type=DataFile.ProgramType.TANF,
        year=2025,
        quarter="Q3",
    )
    ParserErrorFactory(
        file=datafile,
        row_number=42,
        item_number="67",
        field_name="FAMILY_AFFILIATION",
        case_number="11111127148",
        rpt_month_year=202504,
        error_message="FAMILY_AFFILIATION is invalid",
        error_type=ParserErrorCategoryChoices.FIELD_VALUE,
        fields_json={"friendly_name": {"FAMILY_AFFILIATION": "Family Affiliation"}},
    )

    output = ActiveClosedErrorReport(datafile).generate()
    return load_workbook(BytesIO(output.getvalue()))


@pytest.mark.django_db
def test_tanf_error_report_has_readme_sheet_first(active_case_error_report_workbook):
    """The generated workbook should match the prototype sheet order."""
    workbook = active_case_error_report_workbook

    assert workbook.sheetnames == ["Readme", "Critical", "Summary"]


@pytest.mark.django_db
def test_tanf_error_report_data_tabs_start_with_headers(
    active_case_error_report_workbook,
):
    """Data tabs should begin directly with column headings."""
    critical = active_case_error_report_workbook["Critical"]
    summary = active_case_error_report_workbook["Summary"]

    assert [cell.value for cell in critical[1]] == [
        "Case Number",
        "Year",
        "Month",
        "Error Message",
        "Item Number",
        "Item Name",
        "Internal Variable Name",
        "Row Number",
        "Error Type",
    ]
    assert [cell.value for cell in summary[1]] == [
        "Year",
        "Month",
        "Error Message",
        "Item Number",
        "Item Name",
        "Internal Variable Name",
        "Error Type",
        "Number Of Occurrences",
    ]
    assert critical["A2"].value == "11111127148"
    assert summary["A2"].value == "2025"
    assert not critical.merged_cells.ranges
    assert not summary.merged_cells.ranges
    assert not any(critical.row_dimensions[row].height for row in range(1, 5))
    assert not any(summary.row_dimensions[row].height for row in range(1, 5))


@pytest.mark.django_db
def test_tanf_error_report_readme_content_and_links(active_case_error_report_workbook):
    """The README sheet should contain the prototype guidance and links."""
    readme = active_case_error_report_workbook["Readme"]

    assert readme["A1"].value == "Error Report Readme"
    assert "Error Reports are generated for each file you submit" in readme["A2"].value
    assert readme["A4"].value == "Error Report Tabs"
    assert readme["B4"].value == "Description"
    assert readme["A5"].value == "Critical"
    assert readme["A5"].hyperlink.location == "Critical!A1"
    assert readme["A6"].value == "Summary"
    assert readme["A6"].hyperlink.location == "Summary!A1"
    assert readme["A8"].value == "Coding Instructions"
    assert readme["A10"].hyperlink.target == (
        "https://www.acf.hhs.gov/ofa/policy-guidance/"
        "tribal-tanf-data-coding-instructions"
    )
    assert readme["A11"].hyperlink.target == (
        "https://www.acf.hhs.gov/ofa/policy-guidance/acf-ofa-pi-23-04"
    )
    assert readme["A13"].value == "Get More Help"

    expected_help_links = {
        "A14": ("Data File Structure", "data-file-structure"),
        "A15": ("Overview of Error Reports", "overview-of-the-error-report"),
        "A16": ("Interpreting Error Types", "interpreting-error-types"),
        "A17": ("Examples of Common Errors", "examples-of-common-errors"),
    }
    for cell, (label, location) in expected_help_links.items():
        assert readme[cell].value == label
        assert readme[cell].hyperlink.target == KNOWLEDGE_CENTER_URL
        assert readme[cell].hyperlink.location == location


@pytest.mark.django_db
def test_tanf_error_report_readme_layout_and_representative_styles(
    active_case_error_report_workbook,
):
    """The README sheet should preserve key prototype layout details."""
    readme = active_case_error_report_workbook["Readme"]

    assert readme.calculate_dimension() == "A1:B17"
    assert {str(merged) for merged in readme.merged_cells.ranges} == {"A2:B2", "A9:B9"}
    assert readme.column_dimensions["A"].width == pytest.approx(25.29, abs=0.01)
    assert readme.column_dimensions["B"].width == pytest.approx(53.71, abs=0.01)

    expected_heights = {
        1: 18.75,
        2: 66,
        5: 60,
        6: 30,
        9: 36.75,
        14: 45,
        15: 45,
        16: 60,
        17: 45,
    }
    for row, height in expected_heights.items():
        assert readme.row_dimensions[row].height == height

    assert readme["A1"].font.sz == 14
    assert readme["A1"].font.bold is True
    assert readme["A1"].font.underline == "single"
    assert readme["A2"].alignment.wrap_text is True
    assert readme["A9"].alignment.wrap_text is True
    assert readme["A4"].font.bold is True
    assert readme["A4"].border.bottom.style == "thin"
    assert readme["A5"].font.underline == "single"
    assert readme["A5"].font.color.rgb == "FF0070C0"

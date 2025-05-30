"""Tests for DataFiles Application."""
import os
from rest_framework import status
from rest_framework.exceptions import ErrorDetail
from rest_framework.test import APIClient
import pytest
import base64
import openpyxl
from tdpservice.data_files.models import DataFile
from tdpservice.parsers import util
from tdpservice.parsers.factory import ParserFactory
from tdpservice.parsers.models import ParserError
from tdpservice.parsers.test.factories import DataFileSummaryFactory


@pytest.mark.usefixtures('db')
class DataFileAPITestBase:
    """A base test class for tests that interact with the DataFileViewSet.

    Provides several fixtures and methods that are commonly used between tests.
    Intended to simplify creating tests for different user flows.
    """

    root_url = '/v1/data_files/'

    @pytest.fixture
    def user(self):
        """User instance that will be used to log in to the API client.

        This fixture must be overridden in each child test class.
        """
        raise NotImplementedError()

    @pytest.fixture
    def test_datafile(self, stt_user, stt):
        """Fixture for small_incorrect_file_cross_validator."""
        test_datafile = util.create_test_datafile('small_incorrect_file_cross_validator.txt', stt_user, stt)
        test_datafile.year = 2021
        test_datafile.quarter = 'Q1'
        test_datafile.save()
        return test_datafile

    @pytest.fixture
    def test_fra_csv_file(self, stt_user, stt):
        """Fixture for small_incorrect_file_cross_validator."""
        test_datafile = util.create_test_datafile('fra.csv', stt_user, stt,
                                                  DataFile.Section.FRA_WORK_OUTCOME_TANF_EXITERS)
        test_datafile.year = 2024
        test_datafile.quarter = 'Q2'
        test_datafile.save()
        return test_datafile

    @pytest.fixture
    def test_fra_xlsx_file(self, stt_user, stt):
        """Fixture for small_incorrect_file_cross_validator."""
        test_datafile = util.create_test_datafile('fra.xlsx', stt_user, stt,
                                                  DataFile.Section.FRA_WORK_OUTCOME_TANF_EXITERS)
        test_datafile.year = 2024
        test_datafile.quarter = 'Q2'
        test_datafile.save()
        return test_datafile

    @pytest.fixture
    def test_ssp_datafile(self, stt_user, stt):
        """Fixture for small_ssp_section1."""
        df = util.create_test_datafile('small_ssp_section1.txt', stt_user, stt, 'SSP Active Case Data')
        df.year = 2024
        df.quarter = 'Q1'
        return df

    @pytest.fixture
    def api_client(self, api_client, user):
        """Provide an API client that is logged in with the default fixture user."""
        api_client.login(username=user.username, password='test_password')
        return api_client

    @staticmethod
    def login_as(request_user):
        """Provide an API client that is logged in with a provided user."""
        client = APIClient()
        client.login(username=request_user.username, password='test_password')
        return client

    @staticmethod
    def get_data_file_record(data_file_data, version, user):
        """Retrieve a data file record using unique constraints."""
        return DataFile.objects.filter(
            slug=data_file_data["slug"],
            year=data_file_data["year"],
            section=data_file_data["section"],
            version=version,
            user=user,
        )

    @staticmethod
    def assert_data_file_created(response):
        """Assert that the data file was created."""
        assert response.status_code == status.HTTP_201_CREATED

    @staticmethod
    def assert_data_file_error(response):
        """Assert that the data file was created."""
        assert response.status_code == status.HTTP_400_BAD_REQUEST

    @staticmethod
    def assert_data_file_rejected(response):
        """Assert that a given data file submission was rejected."""
        assert response.status_code == status.HTTP_403_FORBIDDEN

    @staticmethod
    def assert_data_file_content_matches(response, data_file_id):
        """Assert that the data file download has the expected content."""
        data_file_file = DataFile.objects.get(id=data_file_id)
        assert b''.join(response.streaming_content) == data_file_file.file.read()

    @staticmethod
    def get_spreadsheet(response):
        """Return error report."""
        decoded_response = base64.b64decode(response.data['xls_report'])

        if os.path.exists('mycls.xlsx'):
            os.remove('mycls.xlsx')

        # write the excel file to disk
        with open('mycls.xlsx', 'wb') as f:
            f.write(decoded_response)

        # read the excel file from disk
        wb = openpyxl.load_workbook('mycls.xlsx')
        return wb

    @staticmethod
    def assert_fra_error_report_file_content_matches_with_friendly_names(response):
        """Assert the error report file contents match expected with friendly names."""
        wb = DataFileAPITestBase.get_spreadsheet(response)
        sheet = wb['Error Report']

        assert sheet.cell(row=2, column=1).value == 202403
        assert sheet.cell(row=2, column=2).value == "*****5891"
        assert sheet.cell(row=2, column=3).value == 7
        assert sheet.cell(row=2, column=4).value == ("Duplicate Social Security Number within a month. "
                                                     "Check that individual Social Security Numbers within a "
                                                     "single exit month are not included more than once. "
                                                     "Social Security Number is a duplicate of the record at "
                                                     "line number 6.")

        assert sheet.cell(row=3, column=1).value == 202301
        assert sheet.cell(row=3, column=2).value == "*****5507"
        assert sheet.cell(row=3, column=3).value == 10
        assert sheet.cell(row=3, column=4).value == ("Social Security Number is not valid. Check that the Social "
                                                     "Security Number is 9 digits, does not contain only zeroes "
                                                     "in any one section, and does not contain dashes or other "
                                                     "punctuation. Enter 999999999 if an individual does not have "
                                                     "an Social Security Number.")

    @staticmethod
    def assert_error_report_tanf_file_content_matches_with_friendly_names(response):
        """Assert the error report file contents match expected with friendly names."""
        wb = DataFileAPITestBase.get_spreadsheet(response)
        critical = wb['Critical']
        summary = wb['Summary']

        COL_ERROR_MESSAGE = 4
        COL_NUM_OCCURRENCES = 8

        assert critical.cell(row=1, column=1).value == "Please refer to the most recent versions of the coding " \
            + "instructions (linked below) when looking up items and allowable values during the data revision process"
        assert critical.cell(row=8, column=COL_ERROR_MESSAGE).value == "No records created."
        assert summary.cell(row=7, column=COL_NUM_OCCURRENCES).value == 1

    @staticmethod
    def assert_error_report_ssp_file_content_matches_with_friendly_names(response):
        """Assert the error report file contents match expected with friendly names."""
        wb = DataFileAPITestBase.get_spreadsheet(response)
        critical = wb['Critical']
        summary = wb['Summary']

        COL_ERROR_MESSAGE = 4
        COL_NUM_OCCURRENCES = 8

        assert critical.cell(row=1, column=1).value == "Please refer to the most recent versions of the coding " \
            + "instructions (linked below) when looking up items and allowable values during the data revision process"
        assert critical.cell(row=7, column=COL_ERROR_MESSAGE).value == ("TRAILER: record length is 15 characters "
                                                                        "but must be 23.")
        assert summary.cell(row=7, column=COL_NUM_OCCURRENCES).value == 3

    @staticmethod
    def assert_error_report_file_content_matches_without_friendly_names(response):
        """Assert the error report file contents match expected without friendly names."""
        wb = DataFileAPITestBase.get_spreadsheet(response)
        critical = wb['Critical']
        summary = wb['Summary']

        COL_ERROR_MESSAGE = 4
        COL_NUM_OCCURRENCES = 8

        assert critical.cell(row=1, column=1).value == "Please refer to the most recent versions of the coding " \
            + "instructions (linked below) when looking up items and allowable values during the data revision process"
        assert critical.cell(row=8, column=COL_ERROR_MESSAGE).value == "No records created."
        assert summary.cell(row=7, column=COL_NUM_OCCURRENCES).value == 1

    @staticmethod
    def assert_data_file_exists(data_file_data, version, user):
        """Confirm that a data file matching the provided data exists in the DB."""
        assert DataFile.objects.filter(
            slug=data_file_data["slug"],
            year=data_file_data["year"],
            section=data_file_data["section"],
            version=version,
        ).exists()

    def post_data_file(self, api_client, data_file_data):
        """Submit a data file with the given data."""
        return api_client.post(
            self.root_url,
            data_file_data,
            format='multipart'
        )

    def post_data_file_fra(self, api_client, data_file_data):
        """Submit a data file with the given data."""
        data_file_data['section'] = 'Secondary School Attainment'
        return api_client.post(
            self.root_url,
            data_file_data,
            format='multipart'
        )

    def get_data_file_files(self, api_client):
        """Submit a data file with the given data."""
        return api_client.get(
            self.root_url
        )

    def get_data_file_file(self, api_client, data_file_id):
        """Get data file meta data."""
        return api_client.get(
            f"{self.root_url}{data_file_id}/"
        )

    def download_file(self, api_client, data_file_id):
        """Stream a file for download."""
        return api_client.get(f"{self.root_url}{data_file_id}/download/")

    def download_error_report_file(self, api_client, data_file_id):
        """Download the ParserError xlsx report."""
        return api_client.get(f"{self.root_url}{data_file_id}/download_error_report/")


class TestDataFileAPIAsOfaAdmin(DataFileAPITestBase):
    """Test DataFileViewSet as an OFA Admin user."""

    @pytest.fixture
    def user(self, ofa_admin, stt):
        """Override the default user with ofa_admin for our tests."""
        return ofa_admin

    def test_get_data_file_file_meta_data(self, api_client, data_file_data, user):
        """Assert the meta data the api provides is as expected."""
        response = self.post_data_file(api_client, data_file_data)
        data_file_id = response.data['id']
        assert DataFile.objects.get(id=data_file_id)
        response = self.get_data_file_file(api_client, data_file_id)

        assert response.data['id'] == data_file_id

        assert str(response.data['user']) == data_file_data['user']

        assert response.data['quarter'] == data_file_data['quarter']
        assert response.data['stt'] == data_file_data['stt']
        assert response.data['year'] == data_file_data['year']

    def test_download_data_file_file(self, api_client, data_file_data, user):
        """Test that the file is transmitted with out errors."""
        response = self.post_data_file(api_client, data_file_data)
        data_file_id = response.data['id']
        response = self.download_file(api_client, data_file_id)

        assert response.status_code == status.HTTP_200_OK
        self.assert_data_file_content_matches(response, data_file_id)

    def test_create_data_file_file_entry(self, api_client, data_file_data, user):
        """Test ability to create data file metadata registry."""
        response = self.post_data_file(api_client, data_file_data)
        self.assert_data_file_created(response)
        self.assert_data_file_exists(data_file_data, 1, user)

    def test_create_data_file_fra_no_feat_flag(self, api_client, data_file_data, user):
        """Test ability to create data file metadata registry."""
        response = self.post_data_file_fra(api_client, data_file_data)
        assert response.data == {'section': [ErrorDetail(string='Section cannot be FRA', code='invalid')]}
        self.assert_data_file_error(response)

    def test_create_data_file_fra_with_feat_flag(self, api_client, csv_data_file, user):
        """Test ability to create data file metadata registry."""
        user.feature_flags = {"fra_reports": True}
        user.save()
        response = self.post_data_file_fra(api_client, csv_data_file)
        self.assert_data_file_created(response)
        self.assert_data_file_exists(csv_data_file, 1, user)

    def test_data_file_file_version_increment(
        self,
        api_client,
        data_file_data,
        other_data_file_data,
        user
    ):
        """Test that data file version numbers incremented."""
        response1 = self.post_data_file(api_client, data_file_data)
        response2 = self.post_data_file(api_client, other_data_file_data)

        self.assert_data_file_created(response1)
        self.assert_data_file_created(response2)

        self.assert_data_file_exists(data_file_data, 1, user)
        self.assert_data_file_exists(other_data_file_data, 2, user)

        file_1 = DataFile.objects.get(id=response1.data['id'])
        file_2 = DataFile.objects.get(id=response2.data['id'])

        assert file_1.s3_versioning_id != file_2.s3_versioning_id


class TestDataFileAPIAsDataAnalyst(DataFileAPITestBase):
    """Test DataFileViewSet as a Data Analyst user."""

    @pytest.fixture
    def user(self, data_analyst):
        """Override the default user with data_analyst for our tests."""
        return data_analyst

    @pytest.fixture
    def dfs(self):
        """Fixture for DataFileSummary."""
        return DataFileSummaryFactory.create()

    def test_data_files_data_analyst_permission(self, api_client, data_file_data, user):
        """Test that a Data Analyst is allowed to add data_files to their own STT."""
        response = self.post_data_file(api_client, data_file_data)
        self.assert_data_file_created(response)
        self.assert_data_file_exists(data_file_data, 1, user)

    def test_data_files_data_analyst_not_allowed(
        self, api_client, data_file_data, user
    ):
        """Test that Data Analysts can't add data_files to STTs other than their own."""
        data_file_data['stt'] = data_file_data['stt'] + 1

        response = self.post_data_file(api_client, data_file_data)
        self.assert_data_file_rejected(response)

    def test_download_data_file_file_for_own_stt(
        self, api_client, data_file_data, user
    ):
        """Test that the file is downloaded as expected for a Data Analyst's set STT."""
        response = self.post_data_file(api_client, data_file_data)
        data_file_id = response.data['id']
        response = self.download_file(api_client, data_file_id)

        assert response.status_code == status.HTTP_200_OK
        self.assert_data_file_content_matches(response, data_file_id)

    @pytest.mark.parametrize("file", [
        ('test_fra_csv_file'),
        ('test_fra_xlsx_file'),
    ])
    def test_download_fra_error_report_file_for_own_stt(self, request, api_client, file, dfs):
        """Test that the fra error report file is downloaded as expected for a Data Analyst's set STT."""
        datafile = request.getfixturevalue(file)
        parser = ParserFactory.get_instance(datafile=datafile, dfs=dfs,
                                            section=datafile.section,
                                            program_type=datafile.prog_type)
        parser.parse_and_validate()

        response = self.download_error_report_file(api_client, datafile.id)

        assert response.status_code == status.HTTP_200_OK
        self.assert_fra_error_report_file_content_matches_with_friendly_names(response)

    def test_download_error_report_file_for_own_stt(
        self, api_client, test_datafile, dfs
    ):
        """Test that the error report file is downloaded as expected for a Data Analyst's set STT."""
        parser = ParserFactory.get_instance(datafile=test_datafile, dfs=dfs,
                                            section=test_datafile.section,
                                            program_type=test_datafile.prog_type)
        parser.parse_and_validate()

        response = self.download_error_report_file(api_client, test_datafile.id)

        assert response.status_code == status.HTTP_200_OK
        self.assert_error_report_tanf_file_content_matches_with_friendly_names(response)

    def test_download_error_report_ssp_file_for_own_stt(self, api_client, test_ssp_datafile, dfs):
        """Test that the error report file for an SSP file is downloaded as expected for a Data Analyst's set STT."""
        parser = ParserFactory.get_instance(datafile=test_ssp_datafile, dfs=dfs,
                                            section=test_ssp_datafile.section,
                                            program_type=test_ssp_datafile.prog_type)
        parser.parse_and_validate()
        response = self.download_error_report_file(api_client, test_ssp_datafile.id)

        assert response.status_code == status.HTTP_200_OK
        self.assert_error_report_ssp_file_content_matches_with_friendly_names(response)

    def test_download_error_report_file_for_own_stt_no_fields_json(
        self, api_client, test_datafile, dfs
    ):
        """Test that the error report file is downloaded as expected when no fields_json is added to ParserErrors."""
        parser = ParserFactory.get_instance(datafile=test_datafile, dfs=dfs,
                                            section=test_datafile.section,
                                            program_type=test_datafile.prog_type)
        parser.parse_and_validate()

        # remove the fields' friendly names for all parser errors
        for error in ParserError.objects.all():
            error.fields_json = None
            error.save()

        response = self.download_error_report_file(api_client, test_datafile.id)

        assert response.status_code == status.HTTP_200_OK
        self.assert_error_report_file_content_matches_without_friendly_names(response)

    def test_download_data_file_file_rejected_for_other_stt(
        self,
        api_client,
        data_file_data,
        other_stt,
        user
    ):
        """Test that the download is rejected when user's STT doesn't match."""
        response = self.post_data_file(api_client, data_file_data)
        data_file_id = response.data['id']

        # Update the STT to something other than the user's
        data_file_file = DataFile.objects.get(id=data_file_id)
        data_file_file.stt = other_stt
        data_file_file.save()

        response = self.download_file(api_client, data_file_id)

        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_download_error_report_rejected_for_other_stt(
        self,
        api_client,
        data_file_data,
        other_stt,
        user
    ):
        """Test that the error report download is rejected when user's STT doesn't match."""
        response = self.post_data_file(api_client, data_file_data)
        data_file_id = response.data['id']

        # Update the STT to something other than the user's
        data_file_file = DataFile.objects.get(id=data_file_id)
        data_file_file.stt = other_stt
        data_file_file.save()

        response = self.download_error_report_file(api_client, data_file_id)

        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_data_files_data_upload_ssp(
        self, api_client, data_file_data,
    ):
        """Test that when Data Analysts upload file with ssp true the section name is updated."""
        data_file_data['ssp'] = True

        response = self.post_data_file(api_client, data_file_data)
        assert response.data['section'] == 'SSP Active Case Data'

    def test_data_file_data_upload_tribe(
        self, api_client, data_file_data, stt
    ):
        """Test that when we upload a file for Tribe the section name is updated."""
        stt.type = 'tribe'
        stt.save()
        response = self.post_data_file(api_client, data_file_data)
        assert 'Tribal Active Case Data' == response.data['section']
        stt.type = ''
        stt.save()

    def test_data_files_data_upload_tanf(
        self, api_client, data_file_data,
    ):
        """Test that when Data Analysts upload file with ssp true the section name is updated."""
        data_file_data['ssp'] = False

        response = self.post_data_file(api_client, data_file_data)
        assert response.data['section'] == 'Active Case Data'


class TestDataFileAPIAsInactiveUser(DataFileAPITestBase):
    """Test DataFileViewSet as an inactive user."""

    @pytest.fixture
    def user(self, inactive_user):
        """Override the default user with inactive_user for our tests."""
        return inactive_user

    def test_data_files_inactive_user_not_allowed(
        self,
        api_client,
        inactive_user,
        data_file_data
    ):
        """Test that an inactive user can't add data_files at all."""
        response = self.post_data_file(api_client, data_file_data)
        self.assert_data_file_rejected(response)


class TestDataFileAsOFARegionalStaff(DataFileAPITestBase):
    """Test DataFileViewSet as a Data Analyst user."""

    @pytest.fixture
    def user(self, regional_user):
        """Override the default user with data_analyst for our tests."""
        return regional_user

    def test_cannot_download_data_file(
        self, api_client, regional_data_file_data, user, data_analyst
    ):
        """Test OFA Regional Staff cannot download datafiles."""
        post_client = self.login_as(data_analyst)
        response = self.post_data_file(post_client, regional_data_file_data)
        data_file_id = response.data['id']
        response = self.download_file(api_client, data_file_id)

        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_can_download_error_report_for_own_region(
        self, api_client, regional_data_file_data, user, data_analyst
    ):
        """Test that OFA Regional Staff can download error reports for data files in their own region."""
        post_client = self.login_as(data_analyst)
        response = self.post_data_file(post_client, regional_data_file_data)
        data_file_id = response.data['id']
        response = self.download_error_report_file(api_client, data_file_id)

        assert response.status_code == status.HTTP_200_OK

    def test_cannot_download_error_report_for_other_region(
        self, api_client, other_regional_data_file_data, user, user_in_other_region
    ):
        """Test that OFA Regional Staff cannot download error reports in regions other than their own."""
        post_client = self.login_as(user_in_other_region)
        response = self.post_data_file(post_client, other_regional_data_file_data)
        data_file_id = response.data['id']
        response = self.download_error_report_file(api_client, data_file_id)

        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_cannot_upload_files(self, api_client, regional_data_file_data, user):
        """Test that OFA Regional Staff cannot create datafiles."""
        response = self.post_data_file(api_client, regional_data_file_data)
        assert response.status_code == status.HTTP_403_FORBIDDEN


def multi_year_data_file_data(user, stt):
    """Return data file data that encompasses multiple years."""
    return [{"original_filename": "data_file.txt",
             "quarter": "Q1",
             "user": user,
             "stt": stt,
             "year": 2020,
             "section": "Active Case Data", },
            {"original_filename": "data_file.txt",
             "quarter": "Q1",
             "user": user,
             "stt": stt,
             "year": 2021,
             "section": "Active Case Data", },
            {"original_filename": "data_file.txt",
             "quarter": "Q1",
             "user": user,
             "stt": stt,
             "year": 2022,
             "section": "Active Case Data", }]


@pytest.mark.django_db
def test_list_data_file_years(api_client, data_analyst):
    """Test list of years for which there exist a data file as a data analyst."""
    user = data_analyst

    data_files = multi_year_data_file_data(user, user.stt)

    DataFile.create_new_version(data_files[0])
    DataFile.create_new_version(data_files[1])
    DataFile.create_new_version(data_files[2])

    api_client.login(username=user.username, password="test_password")

    response = api_client.get("/v1/data_files/years")

    assert response.status_code == status.HTTP_200_OK
    assert response.data == [
        2020,
        2021,
        2022
    ]


@pytest.mark.django_db
def test_list_ofa_admin_data_file_years(api_client, ofa_admin, stt):
    """Test list of years for which there exist a data file as an OFA admin."""
    user = ofa_admin

    data_files = multi_year_data_file_data(user, stt)

    DataFile.create_new_version(data_files[0])
    DataFile.create_new_version(data_files[1])
    DataFile.create_new_version(data_files[2])

    api_client.login(username=user.username, password="test_password")

    response = api_client.get(f"/v1/data_files/years/{stt.id}")

    assert response.status_code == status.HTTP_200_OK
    assert response.data == [
        2020,
        2021,
        2022
    ]


@pytest.mark.django_db
def test_list_ofa_admin_data_file_years_positional_stt(api_client, ofa_admin, stt):
    """Test list year fail for OFA admin when no STT is provided."""
    user = ofa_admin

    data1, data2, data3 = multi_year_data_file_data(user, stt)

    DataFile.create_new_version(data1)
    DataFile.create_new_version(data2)
    DataFile.create_new_version(data3)

    api_client.login(username=user.username, password="test_password")

    response = api_client.get("/v1/data_files/years")

    assert response.status_code == status.HTTP_400_BAD_REQUEST


@pytest.mark.django_db
def test_list_ofa_admin_data_file_years_no_self_stt(
    api_client, ofa_admin_stt_user, stt
):
    """Test OFA Admin with no stt assigned can view list of years."""
    user = ofa_admin_stt_user

    data1, data2, data3 = multi_year_data_file_data(user, stt)

    assert user.stt is None

    DataFile.create_new_version(data1)
    DataFile.create_new_version(data2)
    DataFile.create_new_version(data3)

    api_client.login(username=user.username, password="test_password")

    response = api_client.get(f"/v1/data_files/years/{stt.id}")

    assert response.status_code == status.HTTP_200_OK

    assert response.data == [
        2020,
        2021,
        2022
    ]
